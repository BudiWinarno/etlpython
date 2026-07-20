from flask import Blueprint, render_template, request, send_file, flash, redirect
import os

from config import Config
from database import SessionLocal

from models.template import Template
from models.agent import Agent

from services.generate_service import get_mapping
from services.normalize.factory import NormalizeFactory

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from models.agent_invoice_report import AgentInvoiceReport
import pandas as pd
import math

generate_bp = Blueprint("generate", __name__)

STANDARD_HEADERS = [
    "Nama Agen",
    "Kode Customer",
    "Nama Customer",
    "Alamat Customer",
    "Nomor Telepon/HP Customer",
    "Invoice Nomor Agen",
    "Tanggal Invoice",
    "Tipe Customer",
    "Sales",
    "SKU Kode Agen",
    "Nama SKU",
    "Qty Terjual (Karton)",
    "Qty Terjual (Pcs)",
    "% Diskon 1 (Reguler)",
    "% Diskon 2 (Cash)",
    "% Diskon 3 (DC Fee)",
    "% Diskon 4 (Promo 1)",
    "% Diskon 5 (Promo 2)",
    "% Diskon 6 (Rp)",
    "Quantity Bonus",
    "Rafraksi",
    "Total Invoice Value"
]


@generate_bp.route("/generate")
def index():

    db = SessionLocal()

    templates = db.query(Template).all()

    db.close()

    return render_template(
        "generate/index.html",
        templates=templates
    )


@generate_bp.route("/generate", methods=["POST"])
def generate():

    db = SessionLocal()

    template_id = int(request.form["template_id"])
    
    bulan = int(request.form["bulan"])

    tahun = int(request.form["tahun"])

    file = request.files["file"]

    filepath = os.path.join(
        Config.UPLOAD_FOLDER,
        file.filename
    )

    file.save(filepath)

    # Baca excel
    # template = db.query(Template).filter_by(id=template_id).first()

    # agent = db.query(Agent).filter_by(id=template.agent_id).first()

    # normalizer = NormalizeFactory.get(agent.kode_agent)

    # df = normalizer.normalize(filepath)
    
    # Ambil template
    template = db.query(Template).filter_by(id=template_id).first()

    # Ambil agent
    agent = db.query(Agent).filter_by(id=template.agent_id).first()

    # File sudah dinormalisasi
    df = pd.read_excel(filepath)

    # Ambil mapping
    mapping = get_mapping(template_id)
    
    missing_columns = [
    col for col in mapping.keys()
        if col not in df.columns
    ]

    if missing_columns:
        db.close()

        flash(
            "Template Mapping yang dipilih tidak sesuai dengan file hasil normalisasi.",
            "error"
        )

        return redirect("/generate")

    # Ambil hanya kolom yang dimapping
    df = df[list(mapping.keys())]

    # Rename sesuai standard header
    df = df.rename(columns=mapping)
    
    # ==========================
    # Join Item Agent Mapping
    # ==========================

    from models.item_agent_mapping import ItemAgentMapping

    master = (
        db.query(ItemAgentMapping)
        .filter(
            ItemAgentMapping.agent_id == agent.id,
            ItemAgentMapping.is_active == True
        )
        .all()
    )
    
    if not master:
        db.close()
        flash("Master Kode Agent belum terdaftar. Silakan import Master Item Agent terlebih dahulu.", "error")
        return redirect("/generate")

    master_df = pd.DataFrame([
        {
            "SKU Kode Agen": item.kode_sku_agent,
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama SKU JIM": item.nama_sku_jim,
            "Item Box": item.item_box
        }
        for item in master
    ])
    
    df["SKU Kode Agen"] = (
    df["SKU Kode Agen"]
    .astype(str)
    .str.strip()
    )

    master_df["SKU Kode Agen"] = (
        master_df["SKU Kode Agen"]
        .astype(str)
        .str.strip()
    )
    
    # print("=== Setelah rename ===")
    # print(df.columns.tolist())
    
    print("=== DF ===")
    print(df["SKU Kode Agen"].dtype)
    print(df["SKU Kode Agen"].head())

    print("=== MASTER ===")
    print(master_df["SKU Kode Agen"].dtype)
    print(master_df["SKU Kode Agen"].head())

    print(df[["SKU Kode Agen"]].info())
    print(master_df[["SKU Kode Agen"]].info())

    df = df.merge(
        master_df,
        on="SKU Kode Agen",
        how="left"
    )

    df["Kode SKU JIM"] = df["Kode SKU JIM"].where(df["Kode SKU JIM"].notna(), None)
    df["Nama SKU JIM"] = df["Nama SKU JIM"].where(df["Nama SKU JIM"].notna(), None)

    df["Nama Agen"] = agent.nama_agent

    df["Qty Karton"] = (
        df["Qty Terjual (Pcs)"] /
        df["Item Box"].astype(float)
    )

    df["Qty Karton"] = (
        df["Qty Karton"]
        .fillna(0)
        .apply(lambda x: math.floor(x + 0.5))
    )

    df["Kode SKU JIM"] = df["Kode SKU JIM"].apply(
        lambda x: None if pd.isna(x) else str(x)
    )

    df["Nama SKU JIM"] = df["Nama SKU JIM"].apply(
        lambda x: None if pd.isna(x) else str(x)
    )

    # ==========================
    # DEBUG: cek tipe data Kode SKU JIM / Nama SKU JIM
    # ==========================

    bad_kode = df[
        df["Kode SKU JIM"].apply(lambda x: x is not None and not isinstance(x, str))
    ]

    bad_nama = df[
        df["Nama SKU JIM"].apply(lambda x: x is not None and not isinstance(x, str))
    ]

    if not bad_kode.empty:
        print("=== Baris bermasalah di Kode SKU JIM ===")
        print(bad_kode[["SKU Kode Agen", "Kode SKU JIM"]].to_string())

    if not bad_nama.empty:
        print("=== Baris bermasalah di Nama SKU JIM ===")
        print(bad_nama[["SKU Kode Agen", "Nama SKU JIM"]].to_string())

    print("Total baris:", len(df))
    print("Total bad_kode:", len(bad_kode))
    print("Total bad_nama:", len(bad_nama))

    # Ambil template
    # template = db.query(Template).filter_by(id=template_id).first()

    # Ambil agent
    # agent = db.query(Agent).filter_by(id=template.agent_id).first()
    
    exists = (
    db.query(AgentInvoiceReport)
        .filter(
            AgentInvoiceReport.agent_id == agent.id,
            AgentInvoiceReport.bulan == bulan,
            AgentInvoiceReport.tahun == tahun
        )
        .first()
    )
    
    print(exists)
    
    if exists:

        db.close()

        return "Data bulan dan tahun tersebut sudah ada."
    
    # report = AgentInvoiceReport(
    #     agent_id=1,
    #     bulan=6,
    #     tahun=2026,
    #     nama_agen="TEST",
    #     kode_customer="1",
    #     nama_customer="TEST",
    #     kode_sku_agent="JIMYR067",
    #     kode_sku_jim="DDR356293A",
    #     nama_sku_jim="TEST",
    #     item_box=12,
    #     qty_karton=1
    # )

    # db.add(report)
    # db.commit()

    # return "TEST BERHASIL"

    # Isi Nama Agen

    

    # Tambahkan kolom yang belum ada
    for col in STANDARD_HEADERS:

        if col not in df.columns:

            df[col] = None
    
    df_db = df.copy()
    
    df_export = df[
        [
            "Nama Agen",
            "Kode Customer",
            "Nama Customer",
            "Alamat Customer",
            "Nomor Telepon/HP Customer",
            "Invoice Nomor Agen",
            "Tanggal Invoice",
            "Tipe Customer",
            "Sales",
            "SKU Kode Agen",
            "Nama SKU",
            "Qty Terjual (Karton)",
            "Qty Terjual (Pcs)",
            "% Diskon 1 (Reguler)",
            "% Diskon 2 (Cash)",
            "% Diskon 3 (DC Fee)",
            "% Diskon 4 (Promo 1)",
            "% Diskon 5 (Promo 2)",
            "% Diskon 6 (Rp)",
            "Quantity Bonus",
            "Rafraksi",
            "Total Invoice Value"
        ]
    ]

    # Urutkan kolom
    df = df[STANDARD_HEADERS]
    
    # ==========================
    # Insert ke Database
    # ==========================

    df_db["kode_sku_jim"] = df_db["Kode SKU JIM"].fillna("").astype(str)

    for i, (_, row) in enumerate(df_db.iterrows(), start=1):
        try:
            report = AgentInvoiceReport(
                agent_id=agent.id,
                bulan=bulan,
                tahun=tahun,
                nama_agen=row["Nama Agen"],
                kode_customer=row["Kode Customer"],
                nama_customer=row["Nama Customer"],
                alamat_customer=row["Alamat Customer"],
                nomor_telepon_customer=row["Nomor Telepon/HP Customer"],
                invoice_nomor_agen=row["Invoice Nomor Agen"],
                tanggal_invoice=row["Tanggal Invoice"],
                tipe_customer=row["Tipe Customer"],
                sales=row["Sales"],
                kode_sku_agent=row["SKU Kode Agen"],
                nama_sku=row["Nama SKU"],
                qty_terjual_karton=row["Qty Terjual (Karton)"],
                qty_terjual_pcs=row["Qty Terjual (Pcs)"],
                diskon_1_reguler=row["% Diskon 1 (Reguler)"],
                diskon_2_cash=row["% Diskon 2 (Cash)"],
                diskon_3_dc_fee=row["% Diskon 3 (DC Fee)"],
                diskon_4_promo_1=row["% Diskon 4 (Promo 1)"],
                diskon_5_promo_2=row["% Diskon 5 (Promo 2)"],
                diskon_6_rp=row["% Diskon 6 (Rp)"],
                quantity_bonus=row["Quantity Bonus"],
                rafraksi=row["Rafraksi"],
                total_invoice_value=row["Total Invoice Value"],
                kode_sku_jim=row["Kode SKU JIM"],
                nama_sku_jim=row["Nama SKU JIM"],
                item_box=row["Item Box"],
                qty_karton=row["Qty Karton"]
            )

            db.add(report)
            db.flush()   # insert satu-satu

        except Exception as e:
            print("ERROR BARIS:", i)
            print(row.to_dict())
            raise

    db.commit()

    # sementara jgn insert dulu
    # for _, row in df_db.iterrows():

    #     report = AgentInvoiceReport(

    #         agent_id=agent.id,

    #         bulan=bulan,

    #         tahun=tahun,

    #         nama_agen=row["Nama Agen"],

    #         kode_customer=row["Kode Customer"],

    #         nama_customer=row["Nama Customer"],

    #         alamat_customer=row["Alamat Customer"],

    #         nomor_telepon_customer=row["Nomor Telepon/HP Customer"],

    #         invoice_nomor_agen=row["Invoice Nomor Agen"],

    #         tanggal_invoice=row["Tanggal Invoice"],

    #         tipe_customer=row["Tipe Customer"],

    #         sales=row["Sales"],

    #         kode_sku_agent=row["SKU Kode Agen"],

    #         nama_sku=row["Nama SKU"],

    #         qty_terjual_karton=row["Qty Terjual (Karton)"],

    #         qty_terjual_pcs=row["Qty Terjual (Pcs)"],

    #         diskon_1_reguler=row["% Diskon 1 (Reguler)"],

    #         diskon_2_cash=row["% Diskon 2 (Cash)"],

    #         diskon_3_dc_fee=row["% Diskon 3 (DC Fee)"],

    #         diskon_4_promo_1=row["% Diskon 4 (Promo 1)"],

    #         diskon_5_promo_2=row["% Diskon 5 (Promo 2)"],

    #         diskon_6_rp=row["% Diskon 6 (Rp)"],

    #         quantity_bonus=row["Quantity Bonus"],

    #         rafraksi=row["Rafraksi"],

    #         total_invoice_value=row["Total Invoice Value"],

    #         kode_sku_jim=row["Kode SKU JIM"],

    #         nama_sku_jim=row["Nama SKU JIM"],

    #         item_box=row["Item Box"],

    #         qty_karton=row["Qty Karton"]

    #     )

    #     db.add(report)

        # print(
        #     row["Kode SKU JIM"],
        #     row["Nama SKU JIM"],
        #     row["Item Box"],
        #     row["Qty Karton"]
        # )

        # break

    #     print(AgentInvoiceReport.__table__.columns.keys())

    # db.commit()

    output = os.path.join(
        Config.OUTPUT_FOLDER,
        "Hasil_Generate_Invoice" + file.filename
    )

    df_export.to_excel(
        output,
        index=False
    )

    # df.to_excel(output, index=False)

    # Buka file Excel yang sudah dibuat
    wb = load_workbook(output)
    ws = wb.active

    # Format tanggal
    date_style = NamedStyle(name="date_style")
    date_style.number_format = "mm/dd/yyyy"

    for col in ws.iter_cols():

        for cell in col:

            # Jika nilainya berupa datetime
            if hasattr(cell.value, "year"):
                cell.style = date_style

    # Autofit semua kolom
    for column_cells in ws.columns:

        max_length = 0
        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column].width = max_length + 3

    wb.save(output)

    db.close()

    return send_file(
        output,
        as_attachment=True
    )