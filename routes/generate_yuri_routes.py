from flask import Blueprint
from flask import render_template
from flask import request
from database import SessionLocal
from models.agent import Agent
import os
import pandas as pd
from flask import send_file
from config import Config
from models.item_agent_mapping import ItemAgentMapping
from models.agent_invoice_report import AgentInvoiceReport
from models.agent_stock_report import AgentStockReport
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    NamedStyle,
    Alignment,
    Font,
    PatternFill
)

generate_yuri_bp = Blueprint(
    "generate_yuri",
    __name__
)


# ==========================
# Halaman Generate Yuri
# ==========================

@generate_yuri_bp.route("/generate-yuri")
def index():

    db = SessionLocal()

    agents = (
        db.query(Agent)
        .all()
    )

    db.close()

    return render_template(
        "generate_yuri/index.html",
        agents=agents
    )


# ==========================
# Generate Yuri
# ==========================

@generate_yuri_bp.route("/generate-yuri", methods=["POST"])
def generate():

    db = SessionLocal()

    agent_id = int(request.form["agent_id"])

    bulan = int(request.form["bulan"])

    tahun = int(request.form["tahun"])
    
    agent = (
        db.query(Agent)
        .filter(Agent.id == agent_id)
        .first()
    )

    # ==========================
    # Sheet 1
    # Item Agent Mapping
    # ==========================

    item_mapping = (
        db.query(ItemAgentMapping)
        .filter(
            ItemAgentMapping.agent_id == agent_id,
            ItemAgentMapping.is_active == True
        )
        .all()
    )

    sheet1 = pd.DataFrame([
        {
            "No": i + 1,
            "Item Code": item.kode_sku_jim,
            "Item Name": item.nama_sku_jim,
            "Item / Box": item.item_box,
            "Item Group": item.item_group,
            "Kode SKU Agen": item.kode_sku_agent
        }
        for i, item in enumerate(item_mapping)
    ])

    # ==========================
    # Sheet 2
    # Invoice
    # ==========================

    invoice = (
        db.query(AgentInvoiceReport)
        .filter(
            AgentInvoiceReport.agent_id == agent_id,
            AgentInvoiceReport.bulan == bulan,
            AgentInvoiceReport.tahun == tahun
        )
        .order_by(AgentInvoiceReport.id)
        .all()
    )

    sheet2 = pd.DataFrame([
        {
            "Nama Agen": item.nama_agen,
            "Kode Customer": item.kode_customer,
            "Nama Customer": item.nama_customer,
            "Alamat Customer": item.alamat_customer,
            "Nomor Telepon/HP Customer": item.nomor_telepon_customer,
            "Invoice Nomor Agen": item.invoice_nomor_agen,
            "Tanggal Invoice": item.tanggal_invoice,
            "Tipe Customer": item.tipe_customer,
            "Sales": item.sales,
            "SKU Kode Agen": item.kode_sku_agent,
            "Nama SKU": item.nama_sku,
            "": item.qty_terjual_karton,
            "Qty Terjual (Pcs)": item.qty_terjual_pcs,
            "% Diskon 1 (Reguler)": item.diskon_1_reguler,
            "% Diskon 2 (Cash)": item.diskon_2_cash,
            "% Diskon 3 (DC Fee)": item.diskon_3_dc_fee,
            "% Diskon 4 (Promo 1)": item.diskon_4_promo_1,
            "% Diskon 5 (Promo 2)": item.diskon_5_promo_2,
            "% Diskon 6 (Rp)": item.diskon_6_rp,
            "Quantity Bonus": item.quantity_bonus,
            "Rafraksi": item.rafraksi,
            "Total Invoice Value": item.total_invoice_value
        }
        for item in invoice
    ])

    # ==========================
    # Sheet 3
    # Stock
    # ==========================

    stock = (
        db.query(AgentStockReport)
        .filter(
            AgentStockReport.agent_id == agent_id,
            AgentStockReport.bulan == bulan,
            AgentStockReport.tahun == tahun
        )
        .all()
    )

    sheet3 = pd.DataFrame([
        {
            "KODE": item.kode,
            "Kode SKU Agen": item.kode_sku_agent,
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama JIM": item.nama_sku_jim,
            "QTY (Karton)": item.qty_karton
        }
        for item in stock
    ])

    # ==========================
    # Export Excel
    # ==========================

    # output = os.path.join(
    #     Config.OUTPUT_FOLDER,
    #     f"Template_Yuri_{agent_id}_{bulan}_{tahun}.xlsx"
    # )
    output = os.path.join(
        Config.OUTPUT_FOLDER,
        f"{agent.kode_agent}_{tahun}{bulan:02d}.xlsx"
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        sheet1.to_excel(
            writer,
            sheet_name="Kode Produk JIM",
            index=False
        )

        sheet2.to_excel(
            writer,
            sheet_name="Invoice Agen",
            index=False
        )

        sheet3.to_excel(
            writer,
            sheet_name="Stock Agen",
            index=False
        )
    
    # ==========================
    # Format Excel
    # ==========================

    wb = load_workbook(output)

    # Format tanggal Invoice
    ws = wb["Invoice Agen"]

    date_style = NamedStyle(name="date_style")
    date_style.number_format = "mm/dd/yyyy"

    for cell in ws["G"][1:]:

        if cell.value is not None:

            cell.style = date_style
            
    # ==========================
    # Center Header & Data
    # ==========================

    center = Alignment(
        horizontal="center",
        vertical="center"
    )
    
    # ==========================
    # Style Header
    # ==========================

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_fill = PatternFill(
        fill_type="solid",
        start_color="4472C4",
        end_color="4472C4"
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    for ws in wb.worksheets:

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = center

        # Header
        for cell in ws[1]:

            cell.font = header_font

            cell.fill = header_fill

    # Autofit semua sheet
    for ws in wb.worksheets:

        for column_cells in ws.columns:

            max_length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                except Exception:
                    pass

            ws.column_dimensions[column].width = max_length + 3

    wb.save(output)

    db.close()

    return send_file(
        output,
        as_attachment=True
    )