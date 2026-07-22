from flask import Blueprint
from flask import render_template
from flask import request

from database import SessionLocal
from models.agent import Agent
from models.agent_stock_report import AgentStockReport
import os
import pandas as pd
from flask import send_file
from config import Config
from models.agent_invoice_report import AgentInvoiceReport
import math
from models.cmo_template import CMOTemplate
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from models.insell_report import InsellReport
import numpy as np
from openpyxl.styles import Alignment


generate_cmo_bp = Blueprint(
    "generate_cmo",
    __name__
)


# ==========================
# Halaman Generate CMO
# ==========================

@generate_cmo_bp.route("/generate-cmo")
def index():

    db = SessionLocal()

    agents = (
        db.query(Agent)
        .order_by(Agent.kode_agent)
        .all()
    )

    db.close()

    return render_template(
        "generate_cmo/index.html",
        agents=agents
    )


@generate_cmo_bp.route(
    "/generate-cmo",
    methods=["POST"]
)
def generate():

    db = SessionLocal()

    agent_id = int(request.form["agent_id"])
    bulan = int(request.form["bulan"])
    tahun = int(request.form["tahun"])

    # ==========================
    # Ambil Data Agent
    # ==========================

    agent = (
        db.query(Agent)
        .filter(Agent.id == agent_id)
        .first()
    )

    # ==========================
    # Bulan Sebelumnya
    # ==========================

    if bulan == 1:
        stock_bulan = 12
        stock_tahun = tahun - 1
    else:
        stock_bulan = bulan - 1
        stock_tahun = tahun
        
    sales_bulan = bulan
    sales_tahun = tahun

   # ==========================
    # STOCK
    # ==========================

    stock = (
        db.query(AgentStockReport)
        .filter(
            AgentStockReport.agent_id == agent_id,
            AgentStockReport.bulan == stock_bulan,
            AgentStockReport.tahun == stock_tahun
        )
        .order_by(AgentStockReport.id)
        .all()
    )

    sheet_stock = pd.DataFrame([
        {
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama SKU JIM": item.nama_sku_jim,
            "Qty Karton": item.qty_karton
        }
        for item in stock
    ])
    
    # ==========================
    # Stock Agen bulan ini 
    # ==========================
    stock_current = (
        db.query(AgentStockReport)
        .filter(
            AgentStockReport.agent_id == agent_id,
            AgentStockReport.bulan == bulan,
            AgentStockReport.tahun == tahun
        )
        .order_by(AgentStockReport.id)
        .all()
    )

    sheet_stock_current = pd.DataFrame([
        {
            "Kode SKU JIM": item.kode_sku_jim,
            "Qty Karton": item.qty_karton
        }
        for item in stock_current
    ])
    
    # ==========================
    # INSELL
    # ==========================

    insell = (
        db.query(InsellReport)
        .filter(
            InsellReport.agent_id == agent_id,
            InsellReport.bulan == sales_bulan,
            InsellReport.tahun == sales_tahun
        )
        .all()
    )

    sheet_insell = pd.DataFrame([
        {
            "Kode SKU JIM": item.item_code,
            "Nama SKU JIM": item.item_name,
            "Qty Karton": item.so_qty_karton
        }
        for item in insell
    ])
    
    if not sheet_insell.empty:

        sheet_insell = (
            sheet_insell
            .groupby(
                ["Kode SKU JIM", "Nama SKU JIM"],
                as_index=False
            )
            .agg({
                "Qty Karton": "sum"
            })
        )

    else:

        sheet_insell = pd.DataFrame(
            columns=[
                "Kode SKU JIM",
                "Nama SKU JIM",
                "Qty Karton"
            ]
        )

    # ==========================
    # OUTSELL (1 BULAN)
    # ==========================

    outsell = (
        db.query(AgentInvoiceReport)
        .filter(
            AgentInvoiceReport.agent_id == agent_id,
            AgentInvoiceReport.bulan == sales_bulan,
            AgentInvoiceReport.tahun == sales_tahun
        )
        .all()
    )

    sheet_outsell = pd.DataFrame([
        {
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama SKU JIM": item.nama_sku_jim,
            "Qty PCS": item.qty_terjual_pcs,
            "Item Box": item.item_box
        }
        for item in outsell
    ])

    # if not sheet_outsell.empty:

    #     sheet_outsell = (
    #         sheet_outsell
    #         .groupby(
    #             [
    #                 "Kode SKU JIM",
    #                 "Nama SKU JIM"
    #             ],
    #             as_index=False
    #         )
    #         .agg({
    #             "Qty PCS": "sum",
    #             "Item Box": "first"
    #         })
    #     )

    #     sheet_outsell["Qty Karton"] = (
    #         sheet_outsell["Qty PCS"].astype(float) /
    #         sheet_outsell["Item Box"].astype(float)
    #     ).apply(lambda x: math.floor(x + 0.5))

    #     sheet_outsell = sheet_outsell[
    #         [
    #             "Kode SKU JIM",
    #             "Nama SKU JIM",
    #             "Qty Karton"
    #         ]
    #     ]
    if not sheet_outsell.empty:

        # Skip data rusak (NaN)
        sheet_outsell = sheet_outsell.replace("NaN", pd.NA)

        sheet_outsell["Item Box"] = pd.to_numeric(
            sheet_outsell["Item Box"],
            errors="coerce"
        )

        sheet_outsell = sheet_outsell.dropna(
            subset=["Kode SKU JIM", "Item Box"]
        )

        sheet_outsell = (
            sheet_outsell
            .groupby(
                [
                    "Kode SKU JIM",
                    "Nama SKU JIM"
                ],
                as_index=False
            )
            .agg({
                "Qty PCS": "sum",
                "Item Box": "first"
            })
        )

        hasil = (
            sheet_outsell["Qty PCS"].astype(float)
            / sheet_outsell["Item Box"].astype(float)
        )

        if hasil.isna().any():
            print(f"=== OUTSELL BERMASALAH AGENT {agent.kode_agent} ===")
            print(sheet_outsell[hasil.isna()])

        mask = hasil.notna()

        sheet_outsell = sheet_outsell[mask].copy()
        hasil = hasil[mask]

        sheet_outsell["Qty Karton"] = (
            hasil.apply(lambda x: math.floor(x + 0.5))
        )

    else:

        sheet_outsell = pd.DataFrame(
            columns=[
                "Kode SKU JIM",
                "Nama SKU JIM",
                "Qty Karton"
            ]
        )

    # ==========================
    # AVG OUTSELL (3 BULAN)
    # ==========================

    avg_rows = []

    for i in range(0, 3):

        avg_bulan = sales_bulan  - i
        avg_tahun = sales_tahun

        if avg_bulan <= 0:
            avg_bulan += 12
            avg_tahun -= 1

        data = (
            db.query(AgentInvoiceReport)
            .filter(
                AgentInvoiceReport.agent_id == agent_id,
                AgentInvoiceReport.bulan == avg_bulan,
                AgentInvoiceReport.tahun == avg_tahun
            )
            .all()
        )

        avg_rows.extend(data)
        
        for item in avg_rows:
            if item.item_box is None:
                print("MASUKSINIGANNN")
                print("BULAN :", item.bulan)
                print("INVOICE :", item.invoice_nomor_agen)
                print("SKU AGENT :", item.kode_sku_agent)
                print("SKU JIM :", item.kode_sku_jim)
                print("ITEM BOX :", item.item_box)
                print("PCS :", item.qty_terjual_pcs)

    sheet_avg_outsell = pd.DataFrame([
        {
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama SKU JIM": item.nama_sku_jim,
            "Qty PCS": item.qty_terjual_pcs,
            "Item Box": item.item_box
        }
        for item in avg_rows
    ])
    
    # print("KUYYYYYYYYYY")
    # print(sheet_avg_outsell.iloc[3598])
    # print("KUYYYYYYYYYYANGNGNGNGNGNG")
    # print(avg_rows[3598])
    # print(vars(avg_rows[3598]))
    
    print("Jumlah baris :", len(sheet_avg_outsell))

    print(
        sheet_avg_outsell[
            sheet_avg_outsell["Kode SKU JIM"].isna()
        ]
    )

    print(
        sheet_avg_outsell[
            sheet_avg_outsell["Item Box"].isna()
        ]
    )

    if not sheet_avg_outsell.empty:
        
         # Skip data yang rusak (NaN)
        sheet_avg_outsell = sheet_avg_outsell.replace("NaN", pd.NA)

        sheet_avg_outsell["Item Box"] = pd.to_numeric(
            sheet_avg_outsell["Item Box"],
            errors="coerce"
        )

        sheet_avg_outsell = sheet_avg_outsell.dropna(
            subset=["Kode SKU JIM", "Item Box"]
        )
        
        sheet_avg_outsell = (
            sheet_avg_outsell
            .groupby(
                [
                    "Kode SKU JIM",
                    "Nama SKU JIM"
                ],
                as_index=False
            )
            .agg({
                "Qty PCS": "sum",
                "Item Box": "first"
            })
        )
        
        # ==========================
        # CEK DATA YANG BERMASALAH
        # ==========================
        hasil = (
            sheet_avg_outsell["Qty PCS"].astype(float)
            / sheet_avg_outsell["Item Box"].astype(float)
            / 3
        )

        if hasil.isna().any():
            print(f"=== DATA BERMASALAH AGENT {agent.kode_agent} ===")
            print(sheet_avg_outsell[hasil.isna()])

        mask = hasil.notna()

        sheet_avg_outsell = sheet_avg_outsell[mask].copy()
        hasil = hasil[mask]

        sheet_avg_outsell["Avg Qty Outsell Karton"] = (
            hasil.apply(lambda x: math.floor(x + 0.5))
        )
        
        # ==========================
        # DEBUG
        # ==========================

        print("=== ITEM BOX KOSONG ===")
        print(
            sheet_avg_outsell[
                sheet_avg_outsell["Item Box"].isna()
            ]
        )

        print("=== QTY PCS KOSONG ===")
        print(
            sheet_avg_outsell[
                sheet_avg_outsell["Qty PCS"].isna()
            ]
        )

        # sheet_avg_outsell["Avg Qty Outsell Karton"] = (
        #     (
        #         sheet_avg_outsell["Qty PCS"].astype(float)
        #         / sheet_avg_outsell["Item Box"].astype(float)
        #         / 3
        #     )
        #     .apply(lambda x: math.floor(x + 0.5))
        # )

        sheet_avg_outsell = sheet_avg_outsell[
            [
                "Kode SKU JIM",
                "Nama SKU JIM",
                "Avg Qty Outsell Karton"
            ]
        ]

    else:

        sheet_avg_outsell = pd.DataFrame(
            columns=[
                "Kode SKU JIM",
                "Nama SKU JIM",
                "Avg Qty Outsell Karton"
            ]
        )
        
    # ==========================
    # CMO TEMPLATE
    # ==========================

    templates = (
        db.query(CMOTemplate)
        .filter(
            CMOTemplate.agent_id == agent_id,
            CMOTemplate.is_active == True
        )
        .all()
    )
    
    # tempalte lama
    sheet_template = pd.DataFrame([
        {
            "Kode SKU JIM": item.item_code,
            "Nama SKU JIM": item.item_name,
            "Item Group": item.item_group_name,
            "Customer Name": item.customer_name,
            "Berat (kg)": item.berat,
            "Volume (m3)": item.volume,
            "Item Box": item.item_box,
            "Buffer (Hari)": item.buffer_hari,
            "NKA 1": item.nka_1,
            "NKA 2": item.nka_2,
            "NKA 3": item.nka_3,
            "NKA 4": item.nka_4,
            "Min Stock": item.min_stock,
            "Min Stock NKA 1": item.min_stock_nka_1,
            "Min Stock NKA 2": item.min_stock_nka_2,
            "Order Tambahan": item.order_tambahan
        }
        for item in templates
    ])
    
    # ==========================
    # MERGE STOCK Agen bulan lalu
    # ==========================

    sheet_cmo = sheet_template.merge(
        sheet_stock[
            [
                "Kode SKU JIM",
                "Qty Karton"
            ]
        ],
        on="Kode SKU JIM",
        how="left"
    )
    
    sheet_cmo["Qty Karton"] = (
        sheet_cmo["Qty Karton"]
        .fillna(0)
    )
    
    # ==========================
    # MERGE STOCK Agen bulan ini
    # ==========================
    sheet_cmo = sheet_cmo.merge(
        sheet_stock_current.rename(
            columns={
                "Qty Karton": "Stock Agen Akhir Bulan ini"
            }
        ),
        on="Kode SKU JIM",
        how="left"
    )

    sheet_cmo["Stock Agen Akhir Bulan ini"] = (
        sheet_cmo["Stock Agen Akhir Bulan ini"].fillna(0)
    )
    
    # ==========================
    # MERGE OUTSELL
    # ==========================

    sheet_cmo = sheet_cmo.merge(
        sheet_outsell[
            [
                "Kode SKU JIM",
                "Qty Karton"
            ]
        ].rename(
            columns={
                "Qty Karton": "Outsell Karton"
            }
        ),
        on="Kode SKU JIM",
        how="left"
    )

    sheet_cmo["Outsell Karton"] = (
        sheet_cmo["Outsell Karton"]
        .fillna(0)
    )
    
    # ==========================
    # MERGE AVG OUTSELL
    # ==========================

    sheet_cmo = sheet_cmo.merge(
        sheet_avg_outsell[
            [
                "Kode SKU JIM",
                "Avg Qty Outsell Karton"
            ]
        ],
        on="Kode SKU JIM",
        how="left"
    )

    sheet_cmo["Avg Qty Outsell Karton"] = (
        sheet_cmo["Avg Qty Outsell Karton"]
        .fillna(0)
    )
    
    # ==========================
    # MERGE INSELL
    # ==========================
    sheet_cmo = sheet_cmo.merge(
        sheet_insell[
            [
                "Kode SKU JIM",
                "Qty Karton"
            ]
        ].rename(
            columns={
                "Qty Karton": "Total Sales Qty Insell Carton"
            }
        ),
        on="Kode SKU JIM",
        how="left"
    )

    sheet_cmo["Total Sales Qty Insell Carton"] = (
        sheet_cmo["Total Sales Qty Insell Carton"]
        .fillna(0)
    )
    
    sheet_cmo["Selisih Karton Order Agen (Bulan Berikutnya)"] = ""
    sheet_cmo["CMO"] = ""
    sheet_cmo["TOTAL CMO"] = ""
    sheet_cmo["TOTAL BERAT (kg)"] = ""
    sheet_cmo["TOTAL VOLUME (m3)"] = ""
    
    # ==========================
    # RENAME KOLOM
    # ==========================

    sheet_cmo = sheet_cmo.rename(columns={
        "Kode SKU JIM": "Item Code",
        "Nama SKU JIM": "Item Name",
        "Item Group": "Item Group Name",
        "Berat (kg)": "BERAT (kg)",
        "Volume (m3)": "VOLUME (m3)",
        "Item Box": "Item / Box",
        "Qty Karton": "Stock Agen Akhir Bulan lalu",
        "Outsell Karton": "Total Sales Qty Outsell Carton",
        "Min Stock": "MIN STOK",
        "Min Stock NKA 1": "MIN STOCK (NKA) 1",
        "Min Stock NKA 2": "MIN STOCK (NKA) 2",
        "Order Tambahan": "ORDER TAMBAHAN / PRODUCT PROMOSI"
    })
    
    # ==========================
    # HITUNG CMO
    # ==========================
    
    # sheet_cmo["Stock SAP Akhir Bulan ini"] = (
    #     sheet_cmo["Stock Agen Akhir Bulan lalu"].fillna(0).astype(float)
    #     + sheet_cmo["Total Sales Qty Insell Carton"].fillna(0).astype(float)
    #     - sheet_cmo["Total Sales Qty Outsell Carton"].fillna(0).astype(float)
    # )

    # sheet_cmo["CMO"] = (
    #     sheet_cmo["Total Sales Qty Outsell Carton"].fillna(0).astype(float)
    #     + (
    #         (
    #             sheet_cmo["Buffer (Hari)"].fillna(0).astype(float)
    #             / 30
    #         )
    #         * sheet_cmo["Avg Qty Outsell Karton"].fillna(0).astype(float)
    #     )
    #     - sheet_cmo["Stock Agen Akhir Bulan lalu"].fillna(0).astype(float)
    # )
    
    sheet_cmo["Stock SAP Akhir Bulan ini"] = (
        sheet_cmo["Stock Agen Akhir Bulan lalu"].fillna(0).astype(float)
        + sheet_cmo["Avg Qty Outsell Karton"].fillna(0).astype(float)
        - sheet_cmo["Total Sales Qty Insell Carton"].fillna(0).astype(float)
    )
    
    sheet_cmo["CMO"] = np.where(

    sheet_cmo["MIN STOK"].fillna(0).astype(float) > 0,

        # Jika MIN STOK > 0
        (
            sheet_cmo["MIN STOK"].fillna(0).astype(float)
            + sheet_cmo["Total Sales Qty Outsell Carton"].fillna(0).astype(float)
            - sheet_cmo["Stock Agen Akhir Bulan ini"].fillna(0).astype(float)
        ),

        # Jika MIN STOK <= 0
        (
            sheet_cmo["Total Sales Qty Outsell Carton"].fillna(0).astype(float)
            + (
                sheet_cmo["Buffer (Hari)"].fillna(0).astype(float) / 30
                * sheet_cmo["Avg Qty Outsell Karton"].fillna(0).astype(float)
            )
            - sheet_cmo["Stock Agen Akhir Bulan ini"].fillna(0).astype(float)
        )

    )
    
    sheet_cmo["Selisih Karton Order Agen (Bulan Berikutnya)"] = (
        sheet_cmo["CMO"]
        .apply(lambda x: max(0, math.floor(x + 0.5)))
    )
    
    # 3. Hitung TOTAL CMO
    sheet_cmo["TOTAL CMO"] = (
        sheet_cmo["Selisih Karton Order Agen (Bulan Berikutnya)"].fillna(0).astype(float)
        + sheet_cmo["ORDER TAMBAHAN / PRODUCT PROMOSI"].fillna(0).astype(float)
    )
    
    # Hitung Total Berat
    sheet_cmo["TOTAL BERAT (kg)"] = (
        sheet_cmo["TOTAL CMO"].fillna(0).astype(float)
        * sheet_cmo["BERAT (kg)"].fillna(0).astype(float)
    )

    # Hitung Total Volume
    sheet_cmo["TOTAL VOLUME (m3)"] = (
        sheet_cmo["TOTAL CMO"].fillna(0).astype(float)
        * sheet_cmo["VOLUME (m3)"].fillna(0).astype(float)
    )

    # Jika tidak boleh negatif
    # sheet_cmo["CMO"] = sheet_cmo["CMO"].clip(lower=0)

    # ==========================
    # URUTKAN KOLOM
    # ==========================

    sheet_cmo = sheet_cmo[
        [
            "Item Code",
            "Item Name",
            "Item Group Name",
            "Customer Name",
            "BERAT (kg)",
            "VOLUME (m3)",
            "Item / Box",
            "Stock Agen Akhir Bulan lalu",
            "Buffer (Hari)",
            "Avg Qty Outsell Karton",
            "Total Sales Qty Insell Carton",
            "Total Sales Qty Outsell Carton",
            "NKA 1",
            "NKA 2",
            "NKA 3",
            "NKA 4",
            "MIN STOK",
            "Stock Agen Akhir Bulan ini",
            "Stock SAP Akhir Bulan ini",
            "MIN STOCK (NKA) 1",
            "MIN STOCK (NKA) 2",
            "Selisih Karton Order Agen (Bulan Berikutnya)",
            "CMO",
            "ORDER TAMBAHAN / PRODUCT PROMOSI",
            "TOTAL CMO",
            "TOTAL BERAT (kg)",
            "TOTAL VOLUME (m3)"
        ]
    ]
    
    # sheet_template = pd.DataFrame([
    #     {
    #         "Kode SKU JIM": t.item_code,
    #         "Nama SKU JIM": t.item_name,
    #         "Buffer": t.buffer_hari,
    #         "Berat": t.berat,
    #         "Volume": t.volume,
    #         "Item Box": t.item_box,
    #         "Order Tambahan": t.order_tambahan
    #     }
    #     for t in templates
    # ])

    # ==========================
    # EXPORT EXCEL
    # ==========================

    output = os.path.join(
        Config.OUTPUT_FOLDER,
        f"CMO_{agent.kode_agent}_{tahun}{bulan:02d}.xlsx"
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # sheet_stock.to_excel(
        #     writer,
        #     sheet_name="Stock",
        #     index=False
        # )

        # sheet_outsell.to_excel(
        #     writer,
        #     sheet_name="Outsell",
        #     index=False
        # )

        # sheet_avg_outsell.to_excel(
        #     writer,
        #     sheet_name="Avg Outsell",
        #     index=False
        # )
        
        # sheet_template.to_excel(
        #     writer,
        #     sheet_name="CMO Template",
        #     index=False
        # )
        
        # sheet_cmo.to_excel(
        #     writer,
        #     sheet_name=agent.kode_agent,
        #     index=False
        # )
        
        sheet_cmo.to_excel(
            writer,
            sheet_name=agent.kode_agent,
            index=False,
            startrow=4
        )
        
        ws = writer.sheets[agent.kode_agent]
        
        # ==========================
        # HEADER SUMMARY
        # ==========================

        ws["A1"] = "STOCK AGEN AKHIR BULAN INI"
        ws["B1"] = "STOCK SAP AKHIR BULAN INI"
        ws["C1"] = "AVG OUTSELL"
        ws["D1"] = "OUTSELL"
        ws["E1"] = "INSELL"
        ws["F1"] = "TOTAL CMO"
        ws["G1"] = "TOTAL BERAT"
        ws["H1"] = "TOTAL VOLUME"
        
        # ==========================
        # TOTAL SUMMARY
        # ==========================

        # ws["A2"] = "=SUM(R6:R1048576)"
        # ws["B2"] = "=SUM(S6:S1048576)"
        # ws["C2"] = "=SUM(J6:J1048576)"
        # ws["D2"] = "=SUM(L6:L1048576)"
        # ws["E2"] = "=SUM(K6:K1048576)"
        # ws["F2"] = "=SUM(Y6:Y1048576)"
        # ws["G2"] = "=SUM(Z6:Z1048576)"
        # ws["H2"] = "=SUM(AA6:AA1048576)"
        
        ws["A2"] = sheet_cmo["Stock Agen Akhir Bulan ini"].sum()
        ws["B2"] = sheet_cmo["Stock SAP Akhir Bulan ini"].sum()
        ws["C2"] = sheet_cmo["Avg Qty Outsell Karton"].sum()
        ws["D2"] = sheet_cmo["Total Sales Qty Outsell Carton"].sum()
        ws["E2"] = sheet_cmo["Total Sales Qty Insell Carton"].sum()
        ws["F2"] = sheet_cmo["TOTAL CMO"].sum()
        ws["G2"] = sheet_cmo["TOTAL BERAT (kg)"].sum()
        ws["H2"] = sheet_cmo["TOTAL VOLUME (m3)"].sum()
        
        for row in ws["A1:H2"]:
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # ==========================
        # STYLE HEADER
        # ==========================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"      # Biru tua
        )

        header_font = Font(
            color="FFFFFF",
            bold=True
        )

        thin = Side(style="thin", color="000000")

        header_border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )
        
        # Warna header summary (baris 1)
        for cell in ws["A1:H1"][0]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = header_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ==========================
        # BORDER SEMUA CELL
        # ==========================

        for row in ws.iter_rows():
            for cell in row:
                cell.border = header_border

        # ==========================
        # AUTO FIT
        # ==========================

        for column_cells in ws.columns:
            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )

            ws.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = length + 3

    db.close()

    return send_file(
        output,
        as_attachment=True
    )