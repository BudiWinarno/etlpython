from flask import Blueprint, render_template, request, send_file
import os

from config import Config
from database import SessionLocal

from models.agent import Agent
from models.stock_template import StockTemplate
from models.stock_template_mapping import StockTemplateMapping

from services.normalize.stock.factory import StockNormalizeFactory
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from models.item_agent_mapping import ItemAgentMapping
import pandas as pd
import math
from models.agent_stock_report import AgentStockReport

generate_stock_bp = Blueprint(
    "generate_stock",
    __name__
)

STANDARD_HEADERS = [
    "Kode SKU Agen",
    "QTY (Pcs)"
]


@generate_stock_bp.route("/generate-stock")
def index():

    db = SessionLocal()

    templates = (
        db.query(StockTemplate)
        .filter(StockTemplate.is_active == True)
        .all()
    )
    
    db.close()

    return render_template(
        "generate_stock/index.html",
        templates=templates
    )


@generate_stock_bp.route("/generate-stock", methods=["POST"])
def generate():

    db = SessionLocal()

    template_id = int(
        request.form["template_id"]
    )
    
    bulan = int(request.form["bulan"])

    tahun = int(request.form["tahun"])


    file = request.files["file"]

    filepath = os.path.join(
        Config.UPLOAD_FOLDER,
        file.filename
    )

    file.save(filepath)

    # ==========================
    # Ambil Template
    # ==========================

    template = (
        db.query(StockTemplate)
        .filter_by(id=template_id)
        .first()
    )

    agent = (
        db.query(Agent)
        .filter_by(id=template.agent_id)
        .first()
    )
    
    exists = (
        db.query(AgentStockReport)
        .filter(
            AgentStockReport.agent_id == agent.id,
            AgentStockReport.bulan == bulan,
            AgentStockReport.tahun == tahun
        )
        .first()
    )

    if exists:

        db.close()

        return "Data stock bulan dan tahun tersebut sudah ada."

    # ==========================
    # Normalisasi
    # ==========================

    normalizer = StockNormalizeFactory.get(
        agent.kode_agent
    )

    df = normalizer.normalize(filepath)

    # ==========================
    # Ambil Mapping
    # ==========================

    mappings = (
        db.query(StockTemplateMapping)
        .filter_by(template_id=template_id)
        .all()
    )

    mapping = {}

    for m in mappings:

        mapping[m.excel_header] = m.standard_header

    # ==========================
    # Ambil Kolom
    # ==========================

    df = df[list(mapping.keys())]

    # ==========================
    # Rename Header
    # ==========================

    df = df.rename(columns=mapping)
    
    master = (
        db.query(ItemAgentMapping)
        .filter(
            ItemAgentMapping.agent_id == template.agent_id,
            ItemAgentMapping.is_active == True
        )
        .all()
    )
    
    # print(len(master))
    
    master_df = pd.DataFrame([
        {
            "Kode SKU Agen": item.kode_sku_agent,
            "Kode SKU JIM": item.kode_sku_jim,
            "Nama SKU JIM": item.nama_sku_jim,
            "Item Box": item.item_box,
            "Item Group": item.item_group
        }
        for item in master
    ])
    
    # print(master_df.head())
    
    df = df.merge(
        master_df,
        on="Kode SKU Agen",
        how="left"
    )
    
    # print(df.head())
    
    df["KODE"] = df["Kode SKU Agen"]
    
    df = (
        df.groupby(
            [
                "KODE",
                "Kode SKU Agen",
                "Kode SKU JIM",
                "Nama SKU JIM"
            ],
            as_index=False
        )
        .agg({
            "QTY (Pcs)": "sum",
            "Item Box": "first"
        })
    )
    
    df["QTY (Karton)"] = (
        df["QTY (Pcs)"] / df["Item Box"].astype(float)
    ).apply(lambda x: math.floor(x + 0.5))
    
    # df["KODE"] = df["Kode SKU Agen"]

    # df["Kode SKU JIM"] = df["Kode SKU JIM"]
    
    # print("=== MAPPING ===")
    # print(mapping)

    # print("=== COLUMNS ===")
    # print(df.columns.tolist())
    
    # print(df.head())
    
    # sementara isi KODE sama dengan Kode SKU Agen
    # df["KODE"] = df["Kode SKU Agen"]

    # # sementara isi Kode SKU JIM sama dengan Kode SKU Agen
    # df["Kode SKU JIM"] = df["Kode SKU Agen"]

    # # sementara Nama SKU JIM sama dengan NamaBarang
    # if "NamaBarang" in df.columns:
    #     df["Nama SKU JIM"] = df["NamaBarang"]

    # ==========================
    # Tambah Kolom Yang Belum Ada
    # ==========================

    for col in STANDARD_HEADERS:

        if col not in df.columns:

            df[col] = ""

    # ==========================
    # Urutkan Kolom
    # ==========================

    # df = df[STANDARD_HEADERS]
    df = df[
        [
            "KODE",
            "Kode SKU Agen",
            "Kode SKU JIM",
            "Nama SKU JIM",
            "QTY (Pcs)",
            "Item Box",
            "QTY (Karton)"
        ]
    ]
    
    # ==========================
    # Insert Database
    # ==========================

    for _, row in df.iterrows():

        report = AgentStockReport(

            agent_id=agent.id,

            bulan=bulan,

            tahun=tahun,

            kode=row["KODE"],

            kode_sku_agent=row["Kode SKU Agen"],

            kode_sku_jim=row["Kode SKU JIM"],

            nama_sku_jim=row["Nama SKU JIM"],

            qty_pcs=row["QTY (Pcs)"],

            item_box=row["Item Box"],

            qty_karton=row["QTY (Karton)"]

        )

        db.add(report)

    db.commit()
    
    # ==========================
    # Data Export Excel
    # ==========================

    df_export = df[
        [
            "KODE",
            "Kode SKU Agen",
            "Kode SKU JIM",
            "Nama SKU JIM",
            "QTY (Karton)"
        ]
    ]

    # ==========================
    # Export
    # ==========================

    output = os.path.join(
        Config.OUTPUT_FOLDER,
        "Hasil_Generate_Stock_" + file.filename
    )
    
    df_export.to_excel(
        output,
        index=False
    )

    # ==========================
    # Auto Fit Column
    # ==========================

    wb = load_workbook(output)

    ws = wb.active

    for column_cells in ws.columns:

        max_length = 0

        column = get_column_letter(column_cells[0].column)

        for cell in column_cells:

            try:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            except Exception:
                pass

        ws.column_dimensions[column].width = max_length + 3

    wb.save(output)



    db.close()

    return send_file(
        output,
        as_attachment=True
    )