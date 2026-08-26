import os
import pandas as pd

from openpyxl import load_workbook

from services.normalize.base import BaseNormalizer


class LK000003InvoiceNormalizer(BaseNormalizer):

    def normalize(self, filepath):

        # ==============================
        # Unmerge Cell
        # ==============================

        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(
            ws.merged_cells.ranges
        )

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(
                str(merged)
            )

            for r in range(
                merged.min_row,
                merged.max_row + 1
            ):

                for c in range(
                    merged.min_col,
                    merged.max_col + 1
                ):

                    ws.cell(
                        r,
                        c
                    ).value = value

        temp_file = filepath.replace(
            ".xlsx",
            "_unmerge.xlsx"
        )

        wb.save(temp_file)

        # ==============================
        # Baca Excel
        # ==============================

        df = pd.read_excel(
            temp_file,
            header=None
        )

        # ==============================
        # Debug
        # ==============================

        for i in range(
            min(8, len(df))
        ):

            print(f"ROW {i}")

            print(
                df.iloc[i].tolist()
            )

            print("-" * 80)

        # ==============================
        # Header Utama
        # Row ke-3 Excel
        # index 2 pandas
        # ==============================

        header1 = (
            df.iloc[2]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # ==============================
        # Sub Header
        # Row ke-4 Excel
        # index 3 pandas
        # ==============================

        header2 = (
            df.iloc[3]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # ==============================
        # Gabungkan Header
        # ==============================

        headers = []

        last_header = ""

        for h1, h2 in zip(
            header1,
            header2
        ):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            # Kalau header utama kosong,
            # gunakan header utama sebelumnya

            if h1 == "":
                h1 = last_header

            else:
                last_header = h1

            # Tidak ada sub header

            if h2 == "" or h2.lower() == "nan":

                headers.append(
                    h1
                )

            # Header sama

            elif h1.lower() == h2.lower():

                headers.append(
                    h1
                )

            # Header berbeda

            else:

                headers.append(
                    f"{h1} {h2}"
                )

        # ==============================
        # Mulai Data
        # Row ke-5 Excel
        # index 4 pandas
        # ==============================

        df = (
            df.iloc[4:]
            .reset_index(drop=True)
        )

        df.columns = headers

        # ==============================
        # Rapikan Header
        # ==============================

        new_headers = []

        for col in df.columns:

            col = str(col)

            # Hilangkan spasi berlebih

            col = " ".join(
                col.split()
            )

            # ==========================
            # Normalisasi nama header
            # ==========================

            replacements = {

                "Nama Agen":
                    "Nama Agen",

                "Kode Customer":
                    "Kode Customer",

                "Nama Customer":
                    "Nama Customer",

                "Alamat Customer":
                    "Alamat Customer",

                "Invoice Nomor Agen":
                    "Invoice Nomor Agen",

                "Tanggal Invoice":
                    "Tanggal Invoice",

                "Tipe Customer":
                    "Tipe Customer",

                "Kota":
                    "Kota",

                "SKU Kode Agen":
                    "SKU Kode Agen",

                "Nama SKU":
                    "Nama SKU",

                "Quantity Terjual Karton":
                    "Quantity Terjual Karton",

                "Quantity Terjual PCS":
                    "Quantity Terjual PCS",

                "% Diskon 1 Reguler":
                    "% Diskon 1 Reguler",

                "% Diskon 1 Cash":
                    "% Diskon 1 Cash",

                "% Diskon 1 DC Fee":
                    "% Diskon 1 DC Fee",

                "% Diskon 1 Promo 1":
                    "% Diskon 1 Promo 1",

                "% Diskon 1 Promo 2":
                    "% Diskon 1 Promo 2",

                "Diskon 6 Rp":
                    "Diskon 6 Rp",

                "Quantity Bonus":
                    "Quantity Bonus",

                "Rafraksi (Rp)":
                    "Rafraksi (Rp)",

                "Total Invoice Value":
                    "Total Invoice Value",

                "NAMA SALES":
                    "NAMA SALES",

                "ISI Perkarton":
                    "ISI Perkarton",
            }

            col = replacements.get(
                col,
                col
            )

            col = col.strip()

            new_headers.append(
                col
            )

        df.columns = new_headers

        # ==============================
        # Bersihkan semua header
        # ==============================

        df.columns = (
            df.columns
            .astype(str)
            .str.replace(
                r"\s+",
                " ",
                regex=True
            )
            .str.strip()
        )

        # ==============================
        # Hapus baris kosong
        # ==============================

        df = df.dropna(
            how="all"
        )

        # ==============================
        # Hapus baris header jika ikut
        # terbaca sebagai data
        # ==============================

        if "Nama Agen" in df.columns:

            df = df[
                df["Nama Agen"]
                .astype(str)
                .str.strip()
                .ne("Nama Agen")
            ]

        # ==============================
        # Hapus footer / total
        # ==============================

        if "Invoice Nomor Agen" in df.columns:

            df = df[
                df["Invoice Nomor Agen"]
                .notna()
            ]

        # ==============================
        # Bulatkan Quantity PCS
        # ==============================

        if "Quantity Terjual PCS" in df.columns:

            df["Quantity Terjual PCS"] = (
                pd.to_numeric(
                    df["Quantity Terjual PCS"],
                    errors="coerce"
                )
                .round(0)
            )

        # ==============================
        # Reset Index
        # ==============================

        df = df.reset_index(
            drop=True
        )

        # ==============================
        # Debug
        # ==============================

        print("=" * 80)

        print("COLUMNS:")

        print(
            df.columns.tolist()
        )

        print("\nTOTAL COLUMN:")

        print(
            len(df.columns)
        )

        print("\nTOTAL DATA:")

        print(
            len(df)
        )

        print("\nDATA:")

        print(
            df.head()
        )

        print("=" * 80)

        # ==============================
        # Hapus temporary file
        # ==============================

        os.remove(
            temp_file
        )

        return df

    def get_mapping_headers(self, filepath):

        print("=" * 100)
        print("GET MAPPING HEADERS")
        print("FILEPATH:", filepath)

        # File sudah merupakan hasil normalisasi
        # Jadi header ada di baris pertama
        df = pd.read_excel(
            filepath,
            header=0
        )

        headers = (
            df.columns
            .astype(str)
            .str.replace(
                r"\s+",
                " ",
                regex=True
            )
            .str.strip()
            .tolist()
        )

        print("HASIL MAPPING HEADERS:")
        print(headers)

        print("JUMLAH HEADER:", len(headers))

        print("=" * 100)

        return headers