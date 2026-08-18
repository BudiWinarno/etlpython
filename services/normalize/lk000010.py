import os
import pandas as pd

from openpyxl import load_workbook

from services.normalize.base import BaseNormalizer


class LK000010InvoiceNormalizer(BaseNormalizer):

    def _unmerge_excel(self, filepath, suffix="_unmerge"):
        """
        Unmerge seluruh merged cell dan mengisi semua cell
        dengan value dari cell kiri atas.
        """

        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(ws.merged_cells.ranges)

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(str(merged))

            for r in range(
                merged.min_row,
                merged.max_row + 1
            ):
                for c in range(
                    merged.min_col,
                    merged.max_col + 1
                ):
                    ws.cell(r, c).value = value

        temp_file = filepath.replace(
            ".xlsx",
            f"{suffix}.xlsx"
        )

        wb.save(temp_file)

        return temp_file

    def _build_headers(self, df):

        # Header utama
        header1 = (
            df.iloc[2]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # Sub header
        header2 = (
            df.iloc[3]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        headers = []

        for h1, h2 in zip(header1, header2):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            if h2 == "" or h2.lower() == "nan":
                header = h1

            elif h1 == h2:
                header = h1

            else:
                header = f"{h1} {h2}"

            headers.append(header)

        # Rapikan header
        new_headers = []

        replacements = {
            "Nama Agen Nama Agen": "Nama Agen",
            "Kode Customer Kode Customer": "Kode Customer",
            "Nama Customer Nama Customer": "Nama Customer",
            "Alamat Customer Alamat Customer": "Alamat Customer",
            "Nomor Telepon/HP Customer Nomor Telepon/HP Customer":
                "Nomor Telepon/HP Customer",
            "Invoice Nomor Agen Invoice Nomor Agen":
                "Invoice Nomor Agen",
            "Tanggal Invoice Tanggal Invoice":
                "Tanggal Invoice",
            "Tipe Customer Tipe Customer":
                "Tipe Customer",
            "Kota Kota":
                "Kota",
            "SKU Kode Agen SKU Kode Agen":
                "SKU Kode Agen",
            "Nama SKU Nama SKU":
                "Nama SKU",
            "Quantity Terjual Karton":
                "Quantity Terjual Karton",
            "Quantity Terjual PCS":
                "Quantity Terjual PCS",
            "% Diskon 1 Reguler":
                "% Diskon 1 Reguler",
            "% Diskon 2 Cash":
                "% Diskon 2 Cash",
            "% Diskon 3 DC Fee":
                "% Diskon 3 DC Fee",
            "% Diskon 4 Promo 1":
                "% Diskon 4 Promo 1",
            "% Diskon 5 Promo 2":
                "% Diskon 5 Promo 2",
            "Diskon 6 Rp":
                "Diskon 6 Rp",
            "Quantity Bonus Quantity Bonus":
                "Quantity Bonus",
            "Total Invoice Value Total Invoice Value":
                "Total Invoice Value",
            "Salesman Salesman":
                "Salesman",
        }

        for col in headers:

            col = str(col)

            # Hilangkan spasi berlebih
            col = " ".join(col.split())

            # Rapikan nama tertentu
            col = replacements.get(col, col)

            col = col.strip()

            new_headers.append(col)

        return new_headers

    def normalize(self, filepath):

        # ==================================================
        # UNMERGE CELL
        # ==================================================

        temp_file = self._unmerge_excel(
            filepath,
            "_unmerge"
        )

        # ==================================================
        # BACA EXCEL
        # ==================================================

        df = pd.read_excel(
            temp_file,
            header=None
        )

        # Debug
        print("=" * 80)
        print("PREVIEW EXCEL")

        for i in range(min(8, len(df))):

            print(f"ROW {i}")
            print(df.iloc[i].tolist())
            print("-" * 80)

        # ==================================================
        # BUILD HEADER
        # ==================================================

        headers = self._build_headers(df)

        print("=" * 80)
        print("GENERATED HEADERS")
        print(headers)

        # ==================================================
        # DATA DIMULAI DARI BARIS KE-5 EXCEL
        # pandas index = 4
        # ==================================================

        df = df.iloc[4:].reset_index(drop=True)

        df.columns = headers

        # ==================================================
        # HAPUS BARIS HEADER YANG IKUT TERBACA
        # ==================================================

        if "Nama Agen" in df.columns:

            df = df[
                df["Nama Agen"]
                .astype(str)
                .str.strip()
                .ne("Nama Agen")
            ]

        # ==================================================
        # HAPUS BARIS KOSONG
        # ==================================================

        df = df.dropna(
            how="all"
        ).reset_index(drop=True)

        # ==================================================
        # RAPIIKAN HEADER
        # ==================================================

        df.columns = (
            df.columns
            .astype(str)
            .str.replace(
                r"\s+",
                " ",
                regex=True
            )
            .str.strip()
        )

        # ==================================================
        # RAPIIKAN DATA
        # ==================================================

        for col in df.columns:

            if df[col].dtype == "object":

                df[col] = (
                    df[col]
                    .astype(str)
                    .str.strip()
                )

                # Ubah string nan menjadi None
                df[col] = df[col].replace(
                    {
                        "nan": None,
                        "NaN": None,
                        "None": None,
                    }
                )


            # ==================================================
            # KONVERSI DATA NUMERIC
            # ==================================================

            # Quantity PCS harus integer
            if "Quantity Terjual PCS" in df.columns:

                df["Quantity Terjual PCS"] = (
                    pd.to_numeric(
                        df["Quantity Terjual PCS"],
                        errors="coerce"
                    )
                    .astype("Int64")
                )


            # Quantity Karton bisa desimal
            if "Quantity Terjual Karton" in df.columns:

                df["Quantity Terjual Karton"] = pd.to_numeric(
                    df["Quantity Terjual Karton"],
                    errors="coerce"
                )


            # Quantity Bonus
            if "Quantity Bonus" in df.columns:

                df["Quantity Bonus"] = pd.to_numeric(
                    df["Quantity Bonus"],
                    errors="coerce"
                )


            # Diskon
            numeric_columns = [
                "% Diskon 1 Reguler",
                "% Diskon 2 Cash",
                "% Diskon 3 DC Fee",
                "% Diskon 4 Promo 1",
                "% Diskon 5 Promo 2",
                "Diskon 6 Rp",
                "Total Invoice Value",
            ]

            for col in numeric_columns:

                if col in df.columns:

                    df[col] = pd.to_numeric(
                        df[col],
                        errors="coerce"
                    )


            df = df.reset_index(drop=True)

        # ==================================================
        # DEBUG
        # ==================================================

        print("=" * 80)
        print("FINAL COLUMNS:")
        print(df.columns.tolist())

        print("\nDATA:")
        print(df.head())

        print("\nSHAPE:")
        print(df.shape)

        print("=" * 80)

        # ==================================================
        # HAPUS FILE TEMPORARY
        # ==================================================

        if os.path.exists(temp_file):
            os.remove(temp_file)

        return df

    def get_mapping_headers(self, filepath):

        # ==================================================
        # FILE YANG MASUK KE MAPPING SUDAH HASIL NORMALISASI
        # Jadi langsung baca baris pertama sebagai HEADER
        # ==================================================

        df = pd.read_excel(
            filepath,
            header=0,
            nrows=0
        )

        # ==================================================
        # AMBIL HEADER
        # ==================================================

        headers = (
            df.columns
            .astype(str)
            .str.replace(
                r"\s+",
                " ",
                regex=True
            )
            .str.strip()
            .tolist()
        )

        return headers