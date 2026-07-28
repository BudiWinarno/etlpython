import os
import pandas as pd

from openpyxl import load_workbook

from services.normalize.base import BaseNormalizer

class LK000126InvoiceNormalizer(BaseNormalizer):

    def normalize(self, filepath):

        # ==============================
        # Unmerge Cell
        # ==============================
        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(ws.merged_cells.ranges)

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(str(merged))

            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = value

        temp_file = filepath.replace(".xlsx", "_unmerge.xlsx")

        wb.save(temp_file)

        # ==============================
        # Baca Excel
        # ==============================
        df = pd.read_excel(
            temp_file,
            header=None
        )

        for i in range(8):
            print(f"ROW {i}")
            print(df.iloc[i].tolist())
            print("-" * 80)

        # ==============================
        # Header
        # ==============================
        header1 = (
            df.iloc[1]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        header2 = (
            df.iloc[2]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        headers = []

        for h1, h2 in zip(header1, header2):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            if h2 == "" or h2.lower() == "nan":
                headers.append(h1)

            elif h1 == h2:
                headers.append(h1)

            else:
                headers.append(f"{h1} {h2}")

        # ==============================
        # Mulai Data
        # ==============================
        df = df.iloc[3:].reset_index(drop=True)

        df.columns = headers

        # ==============================
        # Hapus Header Yang Ikut Terbaca
        # ==============================
        df = df[
            df.iloc[:, 0]
            .astype(str)
            .str.strip()
            .ne("Nama Agen")
        ]

        # ==============================
        # Rapikan Header
        # ==============================
        new_headers = []

        replacements = {

            "Nama Agen Nama Agen": "Nama Agen",
            "Kode Customer Kode Customer": "Kode Customer",
            "Nama Customer Nama Customer": "Nama Customer",
            "Alamat Customer Alamat Customer": "Alamat Customer",
            "Invoice Nomor Agen Invoice Nomor Agen": "Invoice Nomor Agen",
            "Tanggal Invoice Tanggal Invoice": "Tanggal Invoice",
            "Tipe Customer Tipe Customer": "Tipe Customer",
            "Kota Kota": "Kota",

            "SKU Kode Agen SKU Kode Agen": "SKU Kode Agen",
            "Nama SKU Nama SKU": "Nama SKU",

            "Quantity Terjual Karton": "Quantity Terjual Karton",
            "Quantity Terjual PCS": "Quantity Terjual PCS",

            "% Diskon 1 Reguler": "% Diskon 1 Reguler",
            "% Diskon 2 Cash": "% Diskon 2 Cash",
            "% Diskon 3 DC Fee": "% Diskon 3 DC Fee",
            "% Diskon 4 Promo 1": "% Diskon 4 Promo 1",
            "% Diskon 5 Promo 2": "% Diskon 5 Promo 2",

            "Diskon 6 Rp": "Diskon 6 Rp",

            "Quantity Bonus Quantity Bonus": "Quantity Bonus",

            "Rafraksi (Rp) Rafraksi (Rp)": "Rafraksi (Rp)",
            "Total Invoice Value Total Invoice Value": "Total Invoice Value",

            "NAMA SALES NAMA SALES": "NAMA SALES",
            "Isi Perkarton Isi Perkarton": "Isi Perkarton",
        }

        for col in df.columns:

            col = " ".join(str(col).split())
            col = replacements.get(col, col)
            new_headers.append(col.strip())

        df.columns = new_headers

        # ==============================
        # Bersihkan Header
        # ==============================
        df.columns = (
            df.columns
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

        df = df.reset_index(drop=True)

        print("=" * 80)
        print("COLUMNS")
        print(df.columns.tolist())

        print("=" * 80)
        print(df.head())

        os.remove(temp_file)

        return df

    def get_mapping_headers(self, filepath):

        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(ws.merged_cells.ranges)

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(str(merged))

            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = value

        temp_file = filepath.replace(".xlsx", "_mapping.xlsx")

        wb.save(temp_file)

        df = pd.read_excel(
            temp_file,
            header=None
        )

        header1 = (
            df.iloc[1]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        header2 = (
            df.iloc[2]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        headers = []

        for h1, h2 in zip(header1, header2):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            if h2 == "" or h2.lower() == "nan":
                headers.append(h1)

            elif h1 == h2:
                headers.append(h1)

            else:
                headers.append(f"{h1} {h2}")

        headers = [
            " ".join(str(h).split()).strip()
            for h in headers
        ]

        replacements = {

            "Nama Agen Nama Agen": "Nama Agen",
            "Kode Customer Kode Customer": "Kode Customer",
            "Nama Customer Nama Customer": "Nama Customer",
            "Alamat Customer Alamat Customer": "Alamat Customer",
            "Invoice Nomor Agen Invoice Nomor Agen": "Invoice Nomor Agen",
            "Tanggal Invoice Tanggal Invoice": "Tanggal Invoice",
            "Tipe Customer Tipe Customer": "Tipe Customer",
            "Kota Kota": "Kota",

            "SKU Kode Agen SKU Kode Agen": "SKU Kode Agen",
            "Nama SKU Nama SKU": "Nama SKU",

            "Quantity Bonus Quantity Bonus": "Quantity Bonus",

            "Rafraksi (Rp) Rafraksi (Rp)": "Rafraksi (Rp)",
            "Total Invoice Value Total Invoice Value": "Total Invoice Value",

            "NAMA SALES NAMA SALES": "NAMA SALES",
            "Isi Perkarton Isi Perkarton": "Isi Perkarton",
        }

        headers = [
            replacements.get(h, h)
            for h in headers
        ]

        os.remove(temp_file)

        return headers