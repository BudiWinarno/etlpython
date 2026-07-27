import os
import pandas as pd

from openpyxl import load_workbook

from services.normalize.base import BaseNormalizer


class LK000021InvoiceNormalizer(BaseNormalizer):

    def normalize(self, filepath):

        # ==============================
        # Unmerge Cell
        # ==============================
        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(ws.merged_cells.ranges)

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(str(merged))

            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = value

        temp_file = filepath.replace(".xlsx", "_unmerge.xlsx")

        wb.save(temp_file)

        # ==============================
        # Baca Excel
        # ==============================
        df = pd.read_excel(
            temp_file,
            header=None
        )
        
        for i in range(8):
            print(f"ROW {i}")
            print(df.iloc[i].tolist())
            print("-" * 80)

        # Header utama
        header1 = (
            df.iloc[1]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # Sub Header
        header2 = (
            df.iloc[2]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        headers = []

        for h1, h2 in zip(header1, header2):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            if h2 == "" or h2.lower() == "nan":
                headers.append(h1)

            elif h1 == h2:
                headers.append(h1)

            else:
                headers.append(f"{h1} {h2}")

        # Mulai data
        df = df.iloc[3:].reset_index(drop=True)

        df.columns = headers

        # ==============================
        # Hapus baris header yang ikut terbaca
        # ==============================
        df = df[
            df.iloc[:, 0]
            .astype(str)
            .str.strip()
            .ne("Nama Agen")
        ]

        # ==============================
        # Rapikan Header
        # ==============================

        new_headers = []

        for col in df.columns:

            col = str(col)

            # Hilangkan spasi berlebih
            col = " ".join(col.split())

            # Rapikan header yang dobel
            replacements = {
                "Nama Agen Nama Agen": "Nama Agen",
                "Kode Customer Kode Customer": "Kode Customer",
                "Nama Customer Nama Customer": "Nama Customer",
                "Alamat Customer Alamat Customer": "Alamat Customer",
                "Invoice Nomor Agen Invoice Nomor Agen": "Invoice Nomor Agen",
                "Tanggal Invoice Tanggal Invoice": "Tanggal Invoice",
                "Tipe Customer Tipe Customer": "Tipe Customer",
                "Kota Kota": "Kota",
                "SKU Kode Agen SKU Kode Agen": "SKU Kode Agen",
                "Nama SKU Nama SKU": "Nama SKU",
                "PCS PCS": "PCS",
                "Rafraksi (Rp) Rafraksi (Rp)": "Rafraksi (Rp)",
                "Total Invoice Value Total Invoice Value": "Total Invoice Value",
            }

            col = replacements.get(col, col)

            # Hilangkan spasi di awal dan akhir
            col = col.strip()

            new_headers.append(col)

        df.columns = new_headers
        
        # Pastikan semua header bersih
        df.columns = (
            df.columns
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
        )

        df = df.reset_index(drop=True)
        
        print("=" * 80)
        print("COLUMNS:")
        print(df.columns.tolist())

        print("\nDATA:")
        print(df.head())

        print("=" * 80)

        os.remove(temp_file)

        return df
    
    def get_mapping_headers(self, filepath):

        wb = load_workbook(filepath)
        ws = wb.active

        merged_ranges = list(ws.merged_cells.ranges)

        for merged in merged_ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            ws.unmerge_cells(str(merged))

            for r in range(merged.min_row, merged.max_row + 1):
                for c in range(merged.min_col, merged.max_col + 1):
                    ws.cell(r, c).value = value

        temp_file = filepath.replace(".xlsx", "_mapping.xlsx")

        wb.save(temp_file)

        df = pd.read_excel(
            temp_file,
            header=None
        )

        headers = (
            df.iloc[0]
            .fillna("")
            .astype(str)
            .str.replace(r"\s+", " ", regex=True)
            .str.strip()
            .tolist()
        )

        os.remove(temp_file)

        return headers