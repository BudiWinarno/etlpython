import os
import pandas as pd

from openpyxl import load_workbook

from services.normalize.base import BaseNormalizer


class LK000035InvoiceNormalizer(BaseNormalizer):

    # =========================================================
    # UNMERGE EXCEL
    # =========================================================
    def _unmerge_excel(self, filepath, suffix):

        temp_file = (
            os.path.splitext(filepath)[0]
            + suffix
            + ".xlsx"
        )

        wb = load_workbook(
            filepath,
            data_only=True
        )

        for ws in wb.worksheets:

            merged_ranges = list(
                ws.merged_cells.ranges
            )

            for merged in merged_ranges:

                value = ws.cell(
                    merged.min_row,
                    merged.min_col
                ).value

                ws.unmerge_cells(
                    str(merged)
                )

                for r in range(
                    merged.min_row,
                    merged.max_row + 1
                ):

                    for c in range(
                        merged.min_col,
                        merged.max_col + 1
                    ):

                        ws.cell(
                            r,
                            c
                        ).value = value

        wb.save(temp_file)

        return temp_file

    # =========================================================
    # NORMALIZE
    # =========================================================
    def normalize(self, filepath):

        # =====================================================
        # 1. UNMERGE
        # =====================================================

        temp_file = self._unmerge_excel(
            filepath,
            "_unmerge"
        )

        # =====================================================
        # 2. BACA EXCEL
        # =====================================================

        df = pd.read_excel(
            temp_file,
            header=None
        )

        # =====================================================
        # 3. HEADER UTAMA
        # =====================================================

        header1 = (
            df.iloc[4]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # =====================================================
        # 4. SUB HEADER
        # =====================================================

        header2 = (
            df.iloc[5]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # =====================================================
        # 5. GABUNG HEADER
        #
        # Kalau ada sub-header:
        #
        # Jul
        # Kuantitas
        #
        # hasil:
        #
        # Kuantitas
        #
        # Kalau tidak ada sub-header:
        #
        # Tgl Faktur
        #
        # hasil:
        #
        # Tgl Faktur
        # =====================================================

        headers = []

        for h1, h2 in zip(
            header1,
            header2
        ):

            h1 = str(h1).strip()
            h2 = str(h2).strip()

            if (
                h2 == ""
                or h2.lower() == "nan"
            ):

                headers.append(h1)

            else:

                # Untuk header seperti:
                # Jul + Kuantitas
                # gunakan header bawah saja

                headers.append(h2)

        # =====================================================
        # 6. DATA MULAI ROW 7
        # =====================================================

        df = df.iloc[6:].reset_index(
            drop=True
        )

        df.columns = headers

        # =====================================================
        # 7. BERSIHKAN HEADER
        # =====================================================

        df.columns = (
            df.columns
            .astype(str)
            .str.replace(
                r"\s+",
                " ",
                regex=True
            )
            .str.strip()
        )

        # =====================================================
        # 8. HAPUS BARIS KOSONG
        # =====================================================

        df = df.dropna(
            how="all"
        )

        # =====================================================
        # 9. FORWARD FILL DATA INVOICE
        # =====================================================

        invoice_columns = [

            "Tgl Faktur",

            "No. Faktur",

            "No. Pelanggan",

            "Nama Pelanggan",

            "Nama Penjual",

            "Alamat 1 Pelanggan",

            "Kota Pelanggan",

            "Nama Tipe Pelanggan Pelanggan",

        ]

        existing_invoice_columns = [

            col
            for col in invoice_columns
            if col in df.columns

        ]

        if existing_invoice_columns:

            df[
                existing_invoice_columns
            ] = (
                df[
                    existing_invoice_columns
                ]
                .ffill()
            )

        # =====================================================
        # 10. HAPUS BARIS TOTAL
        # =====================================================

        df = df[
            ~df.astype(str)
            .apply(
                lambda row:
                row.str.upper()
                .str.contains(
                    r"TOTAL",
                    regex=True,
                    na=False
                ).any(),
                axis=1
            )
        ]

        # =====================================================
        # 11. BERSIHKAN TEXT
        # =====================================================

        text_columns = [

            "No. Faktur",

            "No. Pelanggan",

            "Nama Pelanggan",

            "Nama Penjual",

            "Alamat 1 Pelanggan",

            "Kota Pelanggan",

            "Nama Tipe Pelanggan Pelanggan",

            "No. Barang",

            "Keterangan Barang",

        ]

        for col in text_columns:

            if col in df.columns:

                df[col] = (
                    df[col]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )

        # =====================================================
        # 12. NUMERIC
        # =====================================================

        numeric_columns = [

            "Kuantitas",

            "Jumlah",

            "Harga satuan Barang",

            "Inc ppn 11%",

        ]

        for col in numeric_columns:

            if col in df.columns:

                df[col] = pd.to_numeric(
                    df[col],
                    errors="coerce"
                ).fillna(0)

        # =====================================================
        # 13. DATE
        # =====================================================

        if "Tgl Faktur" in df.columns:

            df["Tgl Faktur"] = pd.to_datetime(
                df["Tgl Faktur"],
                errors="coerce"
            )

        # =====================================================
        # 14. HANYA BARIS YANG ADA BARANG
        # =====================================================

        if (
            "No. Barang" in df.columns
            and
            "Keterangan Barang" in df.columns
        ):

            df = df[
                (
                    df["No. Barang"]
                    .astype(str)
                    .str.strip()
                    .ne("")
                )
                |
                (
                    df["Keterangan Barang"]
                    .astype(str)
                    .str.strip()
                    .ne("")
                )
            ]

        # =====================================================
        # 15. RESET INDEX
        # =====================================================

        df = df.reset_index(
            drop=True
        )

        # =====================================================
        # 16. NOMOR URUT
        # =====================================================

        df.insert(
            0,
            "No",
            range(
                1,
                len(df) + 1
            )
        )

        # =====================================================
        # 17. HAPUS TEMP FILE
        # =====================================================

        if os.path.exists(
            temp_file
        ):

            os.remove(
                temp_file
            )

        # =====================================================
        # 18. RETURN
        # =====================================================

        return df

        # =========================================================
        # GET MAPPING HEADERS
        # =========================================================
        
    def get_mapping_headers(self, filepath):

        # =====================================================
        # BACA FILE HASIL NORMALISASI
        # HEADER SUDAH ADA DI BARIS PERTAMA
        # =====================================================

            df = pd.read_excel(
                filepath,
                header=0,
                nrows=0
            )

            # =====================================================
            # AMBIL HEADER
            # =====================================================

            headers = (
                df.columns
                .astype(str)
                .str.replace(
                    r"\s+",
                    " ",
                    regex=True
                )
                .str.strip()
                .tolist()
            )

            return headers
    